# services/exporter.py — 파사드(총괄)
from __future__ import annotations
import os
from typing import Optional
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# 공통 상수
from .exporter_common import APP_ROOT, TEMPLATES_DIR, EXPORT_DIR

# 다른 내보내기 함수들은 그대로 재노출
from .export_repairs import export_repairs_xlsx
from .export_history_card import export_history_card_xlsx
from .export_consumables import export_consumables_xlsx, save_consumable_template_xlsx
from .export_consumable_txn import export_consumable_txn_xlsx

__all__ = [
    "APP_ROOT","TEMPLATES_DIR","EXPORT_DIR",
    "export_equipment_xlsx","export_repairs_xlsx","export_history_card_xlsx",
    "export_consumables_xlsx","save_consumable_template_xlsx","export_consumable_txn_xlsx",
]

# ─────────────────────────────────────────────────────────────
# 간단한 서식 유틸
def _header(ws, row_idx: int, labels: list[str]):
    ws.append(labels)
    for i in range(1, len(labels)+1):
        c = ws.cell(row=row_idx, column=i)
        c.font = c.font.copy(bold=True)

def _autofit(ws):
    widths = {}
    for row in ws.iter_rows(values_only=True):
        for i, val in enumerate(row, start=1):
            l = 0 if val is None else len(str(val))
            widths[i] = max(widths.get(i, 0), l)
    for i, w in widths.items():
        ws.column_dimensions[get_column_letter(i)].width = min(max(w+2, 8), 60)

def _fmt_date(d):
    return "" if d is None else str(d)

# ─────────────────────────────────────────────────────────────
def export_equipment_xlsx(keyword: str = "", path: Optional[str] = None) -> str:
    """
    설비관리대장 내보내기
    - “용도”는 equipment.purpose
    - “유틸리티 기타”는 equipment.util_other
    """
    from services.equipment_service import list_equipment
    rows = list_equipment(keyword or "")
    wb = Workbook()
    ws = wb.active
    ws.title = "설비관리대장"

    headers = [
        "설비번호", "자산명", "설비명", "설비명 변경안", "모델명",
        "크기(가로x세로x높이)mm", "전압", "전력용량(kWh)",
        "유틸리티 AIR", "유틸리티 냉각수", "유틸리티 진공",
        "용도", "유틸리티 기타",
        "제조회사", "제조회사 대표 전화번호", "제조일자",
        "입고일(년)", "입고일(월)", "입고일(일)",
        "수량", "구입가격", "설비위치", "비고", "파트"
    ]
    _header(ws, 1, headers)

    for e in rows:
        ws.append([
            e.code or "", e.asset_name or "", e.name or "", e.alt_name or "", e.model or "",
            e.size_mm or "", e.voltage or "", e.power_kwh or "",
            e.util_air or "", e.util_coolant or "", e.util_vac or "",
            getattr(e, "purpose", None) or "",       # ★ 용도
            e.util_other or "",                      # ★ 유틸리티 기타
            e.maker or "", e.maker_phone or "", _fmt_date(e.manufacture_date),
            e.in_year or "", e.in_month or "", e.in_day or "",
            e.qty or "", e.purchase_price or "", e.location or "", e.note or "", e.part or ""
        ])

    _autofit(ws)
    if not path:
        os.makedirs(EXPORT_DIR, exist_ok=True)
        path = os.path.join(EXPORT_DIR, "설비관리대장.xlsx")
    else:
        os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    wb.save(path)
    return path
