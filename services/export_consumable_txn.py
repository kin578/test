from __future__ import annotations
import os
from datetime import date, datetime
from typing import Any, List, Optional, Tuple, Dict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from sqlalchemy import text

from db import session_scope
from .exporter_common import EXPORT_DIR, fmt_date

# ─────────────────────────────────────────────────────────────
def _to_date(x) -> Optional[date]:
    """문자열/타입을 date 로 변환. 'YYYY-MM-DD HH:MM:SS'도 지원."""
    if not x:
        return None
    if isinstance(x, date) and not isinstance(x, datetime):
        return x
    if isinstance(x, datetime):
        return x.date()
    if isinstance(x, str):
        s = x.strip()
        if len(s) >= 10 and s[4] == "-" and s[7] == "-":
            try:
                return datetime.strptime(s[:10], "%Y-%m-%d").date()
            except Exception:
                pass
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(s, fmt).date()
            except Exception:
                pass
        try:
            return datetime.fromisoformat(s.replace("Z", "+00:00")).date()
        except Exception:
            return None
    return None

def _best_fetcher():
    """services.consumable_service 안의 조회 함수가 있으면 우선 사용."""
    try:
        from services import consumable_service as cs
    except Exception:
        return None, None
    for name in ("search_consumable_txns", "list_consumable_txns", "list_txns", "get_consumable_txns"):
        fn = getattr(cs, name, None)
        if callable(fn):
            return fn, cs
    return None, None

def _guess_kind(qty: Optional[float], reason: str, related_repair_id: Optional[int]) -> str:
    """구분(입고/출고/수리) 판정."""
    if related_repair_id:
        return "수리"
    if isinstance(qty, (int, float)):
        if qty > 0:
            return "입고"
        if qty < 0:
            return "출고"
    r = (reason or "").strip()
    if "수리" in r:
        return "수리"
    if "출고" in r or "사용" in r or "차감" in r:
        return "출고"
    if "입고" in r or "반입" in r or "추가" in r:
        return "입고"
    return ""

def _normalize_row(row: Dict[str, Any]) -> Tuple[str, str, str, str, float, str, str]:
    """
    표준 스키마로 정규화해 반환:
    (일자, 구분, 품명, 규격, 수량, 사유, 비고)
    """
    dt = row.get("txn_time") or row.get("created_at") or row.get("date")
    reason = str(row.get("reason") or row.get("note") or "")
    qty = float(row.get("qty") or row.get("quantity") or 0)
    name = str(row.get("name") or row.get("consumable_name") or row.get("item_name") or "")
    spec = str(row.get("spec") or row.get("uom") or row.get("unit") or "")
    related_repair_id = row.get("related_repair_id") or row.get("repair_id")

    kind = _guess_kind(qty, reason, related_repair_id)
    qty_out = abs(qty)  # 출고는 구분으로 표현하므로 수량은 절대값

    return (
        fmt_date(_to_date(dt)),
        kind,
        name,
        spec,
        qty_out,
        reason,
        ""  # 비고: 현재 스키마에 별도 없음
    )

def _finalize_path(path: Optional[str], default_name: str) -> str:
    """경로 보정 + 폴더 생성 + 확장자 강제 .xlsx"""
    if not path:
        base = EXPORT_DIR or "."
        os.makedirs(base, exist_ok=True)
        path = os.path.join(base, default_name)
    else:
        path = str(path).strip().strip('"').strip("'")
        if os.path.isdir(path) or path.endswith(("\\", "/")):
            os.makedirs(path, exist_ok=True)
            path = os.path.join(path, default_name)
        root, ext = os.path.splitext(path)
        if not ext or ext.lower() != ".xlsx":
            path = root + ".xlsx"
        os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    return os.path.abspath(path)

# ─────────────────────────────────────────────────────────────
# Fallback: 실제 DB 스키마를 읽어 유연하게 쿼리 구성

def _table_columns(table: str) -> Dict[str, bool]:
    cols: Dict[str, bool] = {}
    with session_scope() as s:
        rows = s.execute(text(f"PRAGMA table_info({table})")).fetchall()
    for r in rows:
        cols[str(r[1]).lower()] = True  # r[1] = column name
    return cols

def _pick(colset: Dict[str, bool], candidates: List[str]) -> Optional[str]:
    for c in candidates:
        if colset.get(c.lower()):
            return c
    return None

def _fallback_query(keyword: str, start_date: Optional[date], end_date: Optional[date]) -> List[dict]:
    """
    네 DB 스키마(consumable_txn/consumable)에 맞춰 안전하게 조회.
    반환 dict에는 txn_time, reason, qty, name, spec, related_repair_id 등이 들어가도록 만든다.
    """
    tcols = _table_columns("consumable_txn")
    ccols = _table_columns("consumable")

    date_col = _pick(tcols, ["txn_time", "created_at", "date"]) or "rowid"
    qty_col  = _pick(tcols, ["qty", "quantity"]) or "qty"
    reason_col = _pick(tcols, ["reason", "note", "memo"])  # 네 DB는 reason
    relrep_col = _pick(tcols, ["related_repair_id", "repair_id"])
    fk_col   = _pick(tcols, ["consumable_id", "consumableid", "cid", "consumable_fk"])

    # 이름/규격은 consumable 테이블에서 가져오고, 없으면 txn의 칼럼을 시도
    join_clause = ""
    if fk_col and ccols:
        name_c = _pick(ccols, ["name", "consumable_name", "item_name"]) or "name"
        spec_c = _pick(ccols, ["spec", "specification"]) or "spec"
        join_clause = f"LEFT JOIN consumable c ON c.id = t.{fk_col}"
        name_expr = f"COALESCE(c.{name_c}, '')"
        spec_expr = f"COALESCE(c.{spec_c}, '')"
    else:
        name_t = _pick(tcols, ["name", "consumable_name", "item_name"])
        spec_t = _pick(tcols, ["spec", "specification", "unit", "uom"])
        name_expr = f"COALESCE(t.{name_t}, '')" if (name_t and tcols.get(name_t.lower())) else "''"
        spec_expr = f"COALESCE(t.{spec_t}, '')" if (spec_t and tcols.get(spec_t.lower())) else "''"

    qty_expr    = f"COALESCE(t.{qty_col}, 0)"
    reason_expr = f"COALESCE(t.{reason_col}, '')" if reason_col else "''"
    relrep_expr = f"t.{relrep_col}" if relrep_col else "NULL"

    # 필터
    conds, params = [], {}
    if date_col != "rowid" and start_date:
        conds.append(f"DATE(t.{date_col}) >= :d1"); params["d1"] = start_date
    if date_col != "rowid" and end_date:
        conds.append(f"DATE(t.{date_col}) <= :d2"); params["d2"] = end_date
    if keyword:
        conds.append(f"({name_expr} LIKE :kw OR {spec_expr} LIKE :kw OR {reason_expr} LIKE :kw)")
        params["kw"] = f"%{keyword}%"
    where_sql = ("WHERE " + " AND ".join(conds)) if conds else ""
    order_col = f"t.{date_col}" if date_col != "rowid" else "t.rowid"

    sql = f"""
        SELECT
            {'t.'+date_col if date_col!='rowid' else 't.rowid'} AS txn_time,
            {reason_expr} AS reason,
            {qty_expr}    AS qty,
            {name_expr}   AS name,
            {spec_expr}   AS spec,
            {relrep_expr} AS related_repair_id
        FROM consumable_txn t
        {join_clause}
        {where_sql}
        ORDER BY {order_col} ASC, t.rowid ASC
    """

    with session_scope() as s:
        rows = s.execute(text(sql), params).mappings().all()
        return [dict(r) for r in rows]

# ─────────────────────────────────────────────────────────────
def export_consumable_txn_xlsx(
    keyword: str = "",
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    path: Optional[str] = None,
    **kwargs,
) -> str:
    """
    호출 예시:
      export_consumable_txn_xlsx()                              # 기본 경로/전체
      export_consumable_txn_xlsx(path="C:/…/a.xlsx")            # 지정 경로
      export_consumable_txn_xlsx(keyword="필터", start_date=…, end_date=…, path=…)
    """
    # (구버전) 첫 인자로 경로만 들어오는 형태 보정
    if path is None and isinstance(keyword, str) and keyword.lower().endswith(".xlsx") and start_date is None and end_date is None:
        path, keyword = keyword, ""

    start_date = _to_date(start_date)
    end_date   = _to_date(end_date)

    # 1) 서비스 함수 우선 사용 (있다면)
    rows: List[Dict[str, Any]] = []
    fetch, _ = _best_fetcher()
    if callable(fetch):
        try:
            rows = list(fetch(keyword or "", start_date, end_date))
        except TypeError:
            try:
                rows = list(fetch(start_date, end_date))
            except TypeError:
                try:
                    rows = list(fetch(keyword or ""))
                except Exception:
                    rows = []
        except Exception:
            rows = []

    # 2) 비면 Fallback(SQL 직조회) + 0건이면 필터없이 재조회
    if not rows:
        rows = _fallback_query(keyword or "", start_date, end_date)
        if not rows and (start_date or end_date):
            rows = _fallback_query(keyword or "", None, None)

    # 엑셀 작성 (요구 컬럼 구성)
    wb = Workbook()
    ws = wb.active
    ws.title = "소모품 입출고"
    headers = ["일자", "구분", "품명", "규격", "수량", "사유", "비고"]
    ws.append(headers)

    for r in rows:
        ws.append(list(_normalize_row(r)))

    # 서식
    widths = [12, 10, 28, 28, 12, 32, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    center = Alignment(horizontal="center", vertical="center")
    right  = Alignment(horizontal="right", vertical="center")
    bold   = Font(bold=True)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c); cell.alignment = center; cell.font = bold
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 1).alignment = center    # 일자
        ws.cell(r, 2).alignment = center    # 구분
        ws.cell(r, 5).alignment = right     # 수량

    # 저장
    out_path = _finalize_path(path, "소모품_입출고이력.xlsx")
    wb.save(out_path)
    if not (os.path.isfile(out_path) and os.path.getsize(out_path) > 0):
        raise RuntimeError(f"파일이 생성되지 않았습니다. 경로/권한을 확인해주세요.\n경로: {out_path}")
    return out_path
