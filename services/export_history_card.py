# services/export_history_card.py
from __future__ import annotations
import os, re, io
from typing import Optional, Dict, List
from datetime import date as _date

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import column_index_from_string
from openpyxl.drawing.image import Image as XLImage
from sqlalchemy import select, and_
from PIL import Image as PILImage

from db import session_scope
from models import Equipment, Repair, Photo
from .exporter_common import (
    fmt_date, safe_sheet_title, ensure_template_history_card,
    find_first_photo_path_for_code, EXPORT_DIR, safe_save_workbook
)

# (선택) 서버 절대경로 해석기 – 없으면 폴백
try:
    from .exporter_common import resolve_photo_abs
except Exception:
    import settings
    APP_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    PHOTOS_FALLBACK_DIR = os.path.join(APP_ROOT, "photos")
    def resolve_photo_abs(rel_or_abs: str | None) -> Optional[str]:
        if not rel_or_abs:
            return None
        if os.path.isabs(rel_or_abs) and os.path.exists(rel_or_abs):
            return rel_or_abs
        root = settings.get_photo_root_dir()
        cand = os.path.join(root, rel_or_abs)
        if os.path.exists(cand):
            return cand
        cand2 = os.path.abspath(os.path.join(APP_ROOT, rel_or_abs))
        if os.path.exists(cand2):
            return cand2
        cand3 = os.path.join(PHOTOS_FALLBACK_DIR, os.path.basename(rel_or_abs))
        if os.path.exists(cand3):
            return cand3
        return None

from services.accessory_service import list_accessories  # 부속기구 목록

# ─────────────────────────────────────────────────────────────
# 유틸
def _norm(s: object) -> str:
    if not isinstance(s, str): return ""
    return re.sub(r"[^0-9a-zA-Z가-힣]", "", s).lower()

def _write_cell(ws: Worksheet, r: int, c: int, val):
    """병합셀 상단좌측 기준으로 안전하게 값 쓰기"""
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= r <= mr.max_row and mr.min_col <= c <= mr.max_col:
            ws.cell(mr.min_row, mr.min_col, val); return
    ws.cell(r, c, val)

def _cell_rc(addr: str):
    import re as _re
    m = _re.match(r"([A-Z]+)(\d+)", addr)
    from openpyxl.utils import column_index_from_string as _cis
    col = _cis(m.group(1)); row = int(m.group(2))
    return row, col

def _anchor_to_rc(anchor):
    try:
        if isinstance(anchor, str):
            import re as _re
            m = _re.match(r"([A-Z]+)(\d+)", anchor)
            if not m: return None
            from openpyxl.utils import column_index_from_string as _cis
            col = _cis(m.group(1)); row = int(m.group(2))
            return (row, col)
        base = getattr(anchor, "_from", None) or getattr(anchor, "from", None)
        if base is not None:
            return (int(getattr(base, "row", 0)) + 1, int(getattr(base, "col", 0)) + 1)
    except Exception:
        pass
    return None

def _wipe_photos_keep_logo(ws: Worksheet, logo_min_row: int = 32):
    """하단 로고(앵커 row >= logo_min_row)만 남기고 사진 삭제"""
    imgs = list(getattr(ws, "_images", []))
    keep = []
    for img in imgs:
        rc = _anchor_to_rc(img.anchor)
        if rc and rc[0] >= logo_min_row:
            keep.append(img)
    ws._images = keep

def _cm_to_px(cm: float, dpi: int = 96) -> int:
    inches = cm / 2.54
    return int(round(inches * dpi))

def _put_image_exact_size(ws: Worksheet, img_path: str, anchor: str, width_cm: float, height_cm: float):
    target_w = _cm_to_px(width_cm); target_h = _cm_to_px(height_cm)
    try:
        with PILImage.open(img_path) as im:
            im = im.resize((target_w, target_h), PILImage.LANCZOS)
            buf = io.BytesIO(); im.save(buf, format="PNG"); buf.seek(0)
        xi = XLImage(buf); xi.width = target_w; xi.height = target_h
        ws.add_image(xi, anchor)
    except Exception:
        pass

# ─────────────────────────────────────────────────────────────
# 수리 이력 좌표
HIST_HEADER_MAP = {
    "년월일": "A27",
    "구분": "D27",
    "고장개소·이력": "E27",
    "조치 내용": "H27",
    "수리처": "J27",
    "수리 시간": "K27",
}
HIST_COL_INDEX = {k: column_index_from_string(v.rstrip("0123456789")) for k, v in HIST_HEADER_MAP.items()}
HIST_START_ROW = 28

def _clear_history_fixed(ws: Worksheet, max_rows: int = 400):
    """템플릿 잔여 이력을 완전히 지움(A~K, 병합 포함)."""
    start = HIST_START_ROW
    end = start + max_rows - 1
    max_col = column_index_from_string("K")

    try:
        ranges = list(ws.merged_cells.ranges)
    except Exception:
        ranges = []
    for mr in ranges:
        if mr.max_row < start or mr.min_col > max_col or mr.min_row == 27:
            continue
        try:
            ws.cell(mr.min_row, mr.min_col).value = None
        except Exception:
            pass

    for r in range(start, end + 1):
        for c in range(1, max_col + 1):
            try:
                _write_cell(ws, r, c, None)
            except Exception:
                pass

# 부속기구 헤더 자동 인식(있으면 채움)
def _norm_str(x): return _norm(x)
def _find_accessory_header(ws: Worksheet):
    wants = [_norm_str(x) for x in ["No", "품명", "규격", "비고"]]
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        vals = [_norm_str(c.value) for c in row]
        hit = sum(1 for w in wants if w in vals)
        if hit >= 3:
            ridx = row[0].row
            cmap = {}
            for c in row:
                v = _norm_str(c.value)
                if v == "no" and "No" not in cmap: cmap["No"] = c.column
                if "품명" in v and "품명" not in cmap: cmap["품명"] = c.column
                if "규격" in v and "규격" not in cmap: cmap["규격"] = c.column
                if "비고" in v and "비고" not in cmap: cmap["비고"] = c.column
            if len(cmap) >= 3:
                return ridx, cmap
    return None

# ─────────────────────────────────────────────────────────────
# 고정 필드 채우기 (용도/특이사항: D13 / A16)
def _fill_fixed_cells(ws: Worksheet, eq: Equipment):
    # 전력/전압 표시
    power = ""
    if getattr(eq, "voltage", None) and getattr(eq, "power_kwh", None) is not None:
        power = f"{eq.voltage}  {eq.power_kwh}kW"
    elif getattr(eq, "voltage", None):
        power = eq.voltage
    elif getattr(eq, "power_kwh", None) is not None:
        power = f"{eq.power_kwh}kW"

    # 입고일 날짜
    in_date = None
    if getattr(eq, "in_year", None):
        y = int(eq.in_year); m = int(getattr(eq, "in_month", 1) or 1); d = int(getattr(eq, "in_day", 1) or 1)
        try:
            from datetime import date as __d
            in_date = __d(y, m, d)
        except Exception:
            in_date = None

    # 기본 셀들
    pairs = {
        "D5":  getattr(eq, "name", "") or "",
        "D6":  getattr(eq, "model", "") or "",
        "D7":  getattr(eq, "size_mm", "") or "",
        "D8":  power,
        "D9":  getattr(eq, "maker", "") or "",
        "D10": fmt_date(in_date),
        "D11": getattr(eq, "purchase_price", None) if getattr(eq, "purchase_price", None) is not None else "",
        "D12": getattr(eq, "location", "") or "",
        # D13은 아래에서 따로(용도) 채움
    }
    for addr, val in pairs.items():
        r, c = _cell_rc(addr); _write_cell(ws, r, c, val)

    # 가격 서식
    r11, c11 = _cell_rc("D11")
    v = ws.cell(r11, c11).value
    try:
        if isinstance(v, str):
            v2 = float(v.replace(",", "").replace(" ", ""))
            ws.cell(r11, c11).value = v2
        ws.cell(r11, c11).number_format = "₩#,##0"
    except Exception:
        ws.cell(r11, c11).number_format = "₩#,##0"

    # TEL (A15)
    tel_text = f"Tel : {getattr(eq, 'maker_phone', '') or ''}"
    r15, c15 = _cell_rc("A15"); _write_cell(ws, r15, c15, tel_text)

    # ★ 정확한 매핑
    # 용도 → D13  (기존 util_other → ❌, 이제 purpose → ⭕)
    purpose = getattr(eq, "purpose", "") or ""
    r_pur, c_pur = _cell_rc("D13")
    _write_cell(ws, r_pur, c_pur, purpose)

    # 특이사항(비고) → A16
    note = getattr(eq, "note", "") or ""
    r_note, c_note = _cell_rc("A16")
    _write_cell(ws, r_note, c_note, note)

def _fill_manager_code_down(ws: Worksheet, code: str):
    # 템플릿의 '관리번호' 라벨 아래로 값 넣기
    for row in ws.iter_rows(min_row=1, max_row=25):
        for cell in row:
            if _norm(str(cell.value)) == _norm("관리번호"):
                _write_cell(ws, cell.row + 1, cell.column, code or "")
                return

# ─────────────────────────────────────────────────────────────
# 워크시트 채우기(연도 필터 지원)
def _fill_sheet_for_code(
    ws: Worksheet,
    equipment_code: str,
    fill_machine_no: bool = False,
    target_year: Optional[int] = None,  # ← 이 연도만 출력(없으면 전체)
):
    with session_scope() as s:
        eq = (
            s.execute(select(Equipment).where(Equipment.code == equipment_code).limit(1))
            .scalars()
            .first()
        )
        if not eq:
            raise ValueError(f"설비({equipment_code})를 찾을 수 없습니다.")

        # shallow copy (세션 분리)
        eq_copy = Equipment()
        for k in ("code","name","model","size_mm","voltage","power_kwh","maker",
                  "in_year","in_month","in_day","purchase_price","location",
                  "purpose","util_other","maker_phone","id","note"):
            setattr(eq_copy, k, getattr(eq, k))
        eq = eq_copy

        # 대표 사진
        photo_rel = s.execute(
            select(Photo.file_path).where(Photo.equipment_id == eq.id).order_by(Photo.id.asc())
        ).scalar()

        # 수리 이력(연도 필터 적용)
        rep_stmt = select(
            Repair.work_date, Repair.kind, Repair.title,
            Repair.detail, Repair.vendor, Repair.work_hours
        ).where(Repair.equipment_id == eq.id)

        if target_year is not None:
            start = _date(target_year, 1, 1)
            end   = _date(target_year, 12, 31)
            rep_stmt = rep_stmt.where(and_(Repair.work_date >= start, Repair.work_date <= end))

        rep_stmt = rep_stmt.order_by(Repair.work_date.asc(), Repair.id.asc())
        reps_raw = s.execute(rep_stmt).all()

    # 고정 필드
    _fill_manager_code_down(ws, eq.code or "")
    _fill_fixed_cells(ws, eq)

    if fill_machine_no:
        # 기기번호 라벨 오른쪽에 코드 출력(있을 때만)
        for row in ws.iter_rows(min_row=1, max_row=25):
            for cell in row:
                if _norm(str(cell.value)) == _norm("기기번호"):
                    _write_cell(ws, cell.row, cell.column + 1, eq.code or "")
                    break

    # 사진(고정 위치/크기)
    _wipe_photos_keep_logo(ws, logo_min_row=32)
    photo_path = resolve_photo_abs(photo_rel) if photo_rel else None
    if not photo_path:
        photo_path = find_first_photo_path_for_code(eq.code or "")
    if photo_path and os.path.isfile(photo_path):
        _put_image_exact_size(ws, photo_path, anchor="G6", width_cm=11.67, height_cm=9.74)

    # 부속기구
    accs = list_accessories(eq.id)
    acc_list = [ (getattr(a, "name", "") or "", getattr(a, "spec", "") or "", getattr(a, "note", "") or "") for a in accs ][:7]
    acc_info = _find_accessory_header(ws)
    if acc_info:
        acc_header_row, acc_col_map = acc_info
        # 초기화
        for i in range(1, 8):
            no_col = acc_col_map.get("No", max(1, acc_col_map.get("품명", 2) - 1))
            _write_cell(ws, acc_header_row + i, no_col, i)
            if "품명" in acc_col_map:  _write_cell(ws, acc_header_row + i, acc_col_map["품명"], "")
            if "규격" in acc_col_map:  _write_cell(ws, acc_header_row + i, acc_col_map["규격"], "")
            if "비고" in acc_col_map:  _write_cell(ws, acc_header_row + i, acc_col_map["비고"], "")
        # 채우기
        for idx, (nm, sp, nt) in enumerate(acc_list, start=1):
            if "품명" in acc_col_map:  _write_cell(ws, acc_header_row + idx, acc_col_map["품명"], nm)
            if "규격" in acc_col_map:  _write_cell(ws, acc_header_row + idx, acc_col_map["규격"], sp)
            if "비고" in acc_col_map:  _write_cell(ws, acc_header_row + idx, acc_col_map["비고"], nt)

    # 수리 이력 표
    _clear_history_fixed(ws, max_rows=400)
    r = HIST_START_ROW
    for wdate, kind, title, detail, vendor, hours in reps_raw:
        _write_cell(ws, r, HIST_COL_INDEX["년월일"], fmt_date(wdate))
        _write_cell(ws, r, HIST_COL_INDEX["구분"], kind or "")
        _write_cell(ws, r, HIST_COL_INDEX["고장개소·이력"], title or "")
        _write_cell(ws, r, HIST_COL_INDEX["조치 내용"], detail or "")
        _write_cell(ws, r, HIST_COL_INDEX["수리처"], vendor or "")
        _write_cell(ws, r, HIST_COL_INDEX["수리 시간"], hours or "")
        r += 1

    ws.title = safe_sheet_title(eq.name or eq.code or "이력카드")

# ─────────────────────────────────────────────────────────────
# 단일/다중 내보내기 (연도 필터 인자 추가)
def export_history_card_xlsx(
    equipment_code: str,
    path: Optional[str] = None,
    template_path: Optional[str] = None,
    logo_path: Optional[str] = None,
    max_repairs: Optional[int] = None,
    fill_machine_no: bool = False,
    year_only: bool = False,              # ← True면 기준일 연도만 출력
    base_date: Optional[_date] = None,    # ← 기준일 (None이면 오늘)
) -> str:
    if not equipment_code:
        raise ValueError("equipment_code가 비어있습니다.")

    tpath = template_path or ensure_template_history_card()
    if tpath and os.path.isfile(tpath):
        wb = load_workbook(tpath); ws = wb.active
    else:
        wb = Workbook(); ws = wb.active; ws.title = "이력카드"

    ty = (base_date or _date.today()).year if year_only else None
    _fill_sheet_for_code(ws, equipment_code, fill_machine_no=fill_machine_no, target_year=ty)

    if not path:
        fn = f"{(equipment_code or 'NONCODE')}_이력카드.xlsx"
        path = os.path.join(EXPORT_DIR, fn)
    return safe_save_workbook(wb, path)

def _unique_title(base: str, used: set[str]) -> str:
    name = safe_sheet_title(base or "Sheet")
    if name not in used:
        used.add(name); return name
    i = 2
    while True:
        cand = safe_sheet_title(f"{name} ({i})")
        if cand not in used:
            used.add(cand); return cand
        i += 1

def export_history_cards_multi_xlsx(
    equipment_codes: List[str],
    path: Optional[str] = None,
    template_path: Optional[str] = None,
    fill_machine_no: bool = False,
    sort_by: str = "code",
    sheet_title_format: Optional[str] = None,
    year_only: bool = False,              # ← True면 기준일 연도만 출력
    base_date: Optional[_date] = None,    # ← 기준일 (None이면 오늘)
) -> str:
    codes = [c for c in (equipment_codes or []) if c]
    if not codes:
        raise ValueError("equipment_codes가 비어있습니다.")

    # 코드→이름 조회(시트명 정렬용)
    name_map: Dict[str, str] = {}
    with session_scope() as s:
        rows = s.execute(select(Equipment.code, Equipment.name).where(Equipment.code.in_(codes))).all()
        for code, name in rows:
            name_map[code] = name or ""

    if sort_by == "name":
        codes.sort(key=lambda c: (name_map.get(c, "") or "", c))
    else:
        codes.sort(key=lambda c: c)

    tpath = template_path or ensure_template_history_card()
    if not (tpath and os.path.isfile(tpath)):
        wb = Workbook(); ws_master = wb.active; ws_master.title = "이력카드"
    else:
        wb = load_workbook(tpath); ws_master = wb.active

    used_titles: set[str] = set()
    ty = (base_date or _date.today()).year if year_only else None

    for idx, code in enumerate(codes):
        ws = ws_master if idx == 0 else wb.copy_worksheet(ws_master)
        _fill_sheet_for_code(ws, code, fill_machine_no=fill_machine_no, target_year=ty)

        nm = name_map.get(code, "") or ""
        base = (sheet_title_format.format(code=code, name=nm) if sheet_title_format
                else (nm or code))
        # 고유 시트명
        name = safe_sheet_title(base or "Sheet")
        if name in used_titles:
            i = 2
            while f"{name} ({i})" in used_titles:
                i += 1
            name = f"{name} ({i})"
        used_titles.add(name)
        ws.title = name

    if not path:
        path = os.path.join(EXPORT_DIR, f"이력카드_묶음_{len(codes)}대.xlsx")
    return safe_save_workbook(wb, path)
