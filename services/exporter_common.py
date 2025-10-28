from __future__ import annotations
import os, io, sys
from datetime import datetime, date
from typing import Optional, Iterable, Dict

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image as PILImage

from sqlalchemy import select
from db import session_scope
from models import Photo

import settings  # 서버/로컬 사진 루트

# (있으면 사용, 없어도 동작)
try:
    from services.photo_service import PHOTO_ROOT  # type: ignore
except Exception:
    PHOTO_ROOT = settings.get_photo_root_dir()

# ─────────────────────────────────────────────
# 배포본은 exe 폴더, 개발은 프로젝트 루트
if getattr(sys, "frozen", False):
    APP_ROOT = os.path.dirname(sys.executable)
else:
    APP_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))

# ✅ PyInstaller 6.x onedir 레이아웃 보강: _internal 경로
INTERNAL_DIR = os.path.join(APP_ROOT, "_internal")

TEMPLATES_DIR = os.path.join(APP_ROOT, "templates")
EXPORT_DIR = os.path.join(APP_ROOT, "exports")
PHOTOS_FALLBACK_DIR = os.path.join(APP_ROOT, "photos")
os.makedirs(TEMPLATES_DIR, exist_ok=True)
os.makedirs(EXPORT_DIR, exist_ok=True)

# ─────────────────────────────────────────────
# 템플릿 탐색 유틸 (우선순위: exe옆 → _internal → CWD → 루트 직하)
def _template_dirs() -> list[str]:
    return [
        os.path.join(APP_ROOT, "templates"),
        os.path.join(INTERNAL_DIR, "templates"),  # ★ 여기 추가
        os.path.join(os.getcwd(), "templates"),
        APP_ROOT,  # (혹시 루트에 파일을 둘 때)
    ]

def get_template_path(filename: str) -> str:
    for d in _template_dirs():
        p = os.path.join(d, filename)
        if os.path.isfile(p):
            return p
    raise FileNotFoundError(
        f"템플릿을 찾지 못했습니다: {filename}\n검색 경로: {_template_dirs()}"
    )

def ensure_template_history_card() -> Optional[str]:
    # 이력카드 샘플 템플릿을 위 우선순위대로 검색
    try:
        return get_template_path("이력카드 샘플.xlsx")
    except FileNotFoundError:
        return None

# ─────────────────────────────────────────────
def fmt_date(d) -> str:
    if not d: return ""
    if isinstance(d, datetime): return d.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(d, date): return d.strftime("%Y-%m-%d")
    try: return str(d)
    except Exception: return ""

def safe_sheet_title(name: str) -> str:
    if not name: name = "Sheet"
    bad = set(r'[]:*?/\\')
    title = "".join(ch for ch in name if ch not in bad)[:31]
    return title or "Sheet"

def autofit(ws: Worksheet, max_width: int = 60):
    lens: Dict[int, int] = {}
    for r in ws.iter_rows(values_only=True):
        for i, v in enumerate(r, start=1):
            s = "" if v is None else str(v)
            est = int(len(s.encode("utf-8")) * 0.6)
            lens[i] = max(lens.get(i, 10), min(est + 2, max_width))
    for i, w in lens.items():
        ws.column_dimensions[get_column_letter(i)].width = max(10, w)

def header(ws: Worksheet, row: int, labels: Iterable[str]):
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    bold = Font(bold=True); fill = PatternFill("solid", fgColor="F2F2F2")
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for i, h in enumerate(labels, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = bold; c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

# ─────────────────────────────────────────────
# 사진 경로/삽입
def resolve_photo_abs(rel_or_abs: str | None) -> Optional[str]:
    if not rel_or_abs: return None
    if os.path.isabs(rel_or_abs) and os.path.exists(rel_or_abs):
        return rel_or_abs
    root = settings.get_photo_root_dir()
    cand = os.path.join(root, rel_or_abs)
    if os.path.exists(cand): return cand
    cand2 = os.path.abspath(os.path.join(APP_ROOT, rel_or_abs))
    if os.path.exists(cand2): return cand2
    cand3 = os.path.join(PHOTOS_FALLBACK_DIR, os.path.basename(rel_or_abs))
    if os.path.exists(cand3): return cand3
    return None

def find_first_photo_path_for_code(code: str) -> Optional[str]:
    if not code: return None
    with session_scope() as s:
        p = None
        try:
            from sqlalchemy import or_
            p = s.execute(
                select(Photo).where(or_(Photo.equipment_code == code,
                                        Photo.file_path.like(f"{code}/%")))
            ).scalars().first()
        except Exception:
            p = None
        if p:
            rel = getattr(p, "file_path", None) or getattr(p, "path", None)
            abs_path = resolve_photo_abs(rel)
            if abs_path and os.path.isfile(abs_path):
                return abs_path

    folder = os.path.join(settings.get_photo_root_dir(), code)
    if os.path.isdir(folder):
        for fn in sorted(os.listdir(folder)):
            if fn.lower().endswith((".png", ".jpg", ".jpeg", ".bmp", ".gif", ".webp")):
                return os.path.join(folder, fn)

    legacy = os.path.join(PHOTOS_FALLBACK_DIR, code)
    if os.path.isdir(legacy):
        for fn in sorted(os.listdir(legacy)):
            if fn.lower().endswith((".png", ".jpg", ".jpeg", ".bmp", ".gif", ".webp")):
                return os.path.join(legacy, fn)
    return None

def put_image(ws: Worksheet, img_path: str, anchor: str, max_w_px: int, max_h_px: int):
    try:
        with PILImage.open(img_path) as im:
            w, h = im.size
            scale = min(max_w_px / max(1, w), max_h_px / max(1, h), 1.0)
            nw, nh = int(w * scale), int(h * scale)
            im = im.resize((nw, nh), PILImage.LANCZOS)
            buf = io.BytesIO(); im.save(buf, format="PNG"); buf.seek(0)
        ox = XLImage(buf); ox.width = nw; ox.height = nh
        ws.add_image(ox, anchor)
    except Exception:
        pass

# ─────────────────────────────────────────────
# 안전 저장
def _ensure_parent_dir(path: str):
    parent = os.path.dirname(path)
    if parent and not os.path.isdir(parent):
        os.makedirs(parent, exist_ok=True)

def _unique_in_dir(dirpath: str, name: str, ext: str) -> str:
    for i in range(1, 100):
        cand = os.path.join(dirpath, f"{name}({i}){ext}")
        if not os.path.exists(cand): return cand
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(dirpath, f"{name}_{ts}{ext}")

def safe_save_workbook(wb: Workbook, path: str) -> str:
    try:
        _ensure_parent_dir(path); wb.save(path); return path
    except PermissionError:
        try:
            d, fn = os.path.split(path); name, ext = os.path.splitext(fn)
            cand = _unique_in_dir(d or ".", name, ext or ".xlsx")
            _ensure_parent_dir(cand); wb.save(cand); return cand
        except Exception:
            name = name if 'name' in locals() else "내보내기"
            ext = ext if 'ext' in locals() and ext else ".xlsx"
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            fallback = os.path.join(EXPORT_DIR, f"{name}_{ts}{ext}")
            _ensure_parent_dir(fallback); wb.save(fallback); return fallback
    except OSError:
        d, fn = os.path.split(path); name, ext = os.path.splitext(fn or "내보내기.xlsx")
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        fallback = os.path.join(EXPORT_DIR, f"{name}_{ts}{ext or '.xlsx'}")
        _ensure_parent_dir(fallback); wb.save(fallback); return fallback
